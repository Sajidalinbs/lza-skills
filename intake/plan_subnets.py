#!/usr/bin/env python3
"""plan_subnets.py — turn a lightweight network *requirement* into a finalized,
overlap-safe CIDR plan and a review-able Excel workbook.

The intake philosophy (per the engagement workflow):
  * The customer states, per VPC, the SUBNET NAMES and HOW MANY IPs each needs.
  * This tool DECIDES the actual CIDRs — sizing each subnet to fit its IP count,
    carving them out of the VPC CIDR, and AVOIDING overlap with on-prem / peer
    networks the customer told us to stay clear of.
  * It writes an Excel workbook so the customer can review the finalized IP plan
    BEFORE any LZA YAML is generated.

Input: a small YAML requirements file (see requirements.example.yaml).
Output:
  * <customer>-network-plan.xlsx   — review artifact (Accounts/OUs/VPCs/Subnets/External-CIDRs)
  * network-config.snippet.yaml    — the vpcs[].subnets[] block for LZA network-config.yaml

Usage:
  python3 plan_subnets.py requirements.example.yaml
  python3 plan_subnets.py requirements.example.yaml --out-dir ./out

AWS facts baked in:
  * 5 IPs are reserved per subnet (network, VPC router, DNS, future, broadcast),
    so a subnet of prefix /p yields 2^(32-p) - 5 usable addresses.
  * Subnet prefixes are clamped to /28 (min size) .. the VPC prefix (max size).
"""
import argparse
import csv
import ipaddress
import os
import sys

try:
    import yaml
except ImportError:
    sys.exit("pyyaml required: pip install pyyaml")
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl required: pip install openpyxl")

AWS_RESERVED = 5
MIN_SUBNET_PREFIX = 28          # smallest subnet AWS allows
MAX_SUBNET_PREFIX = 16          # largest subnet block we'll hand out


# ----- core CIDR math -------------------------------------------------------

def prefix_for_ips(ips_needed, vpc_prefix):
    """Smallest subnet (largest prefix number) whose usable host count >= ips_needed."""
    need = ips_needed + AWS_RESERVED
    floor_prefix = max(MAX_SUBNET_PREFIX, vpc_prefix)  # subnet can't be bigger than its VPC
    for p in range(MIN_SUBNET_PREFIX, floor_prefix - 1, -1):
        if (1 << (32 - p)) >= need:
            return p
    raise ValueError(f"{ips_needed} IPs do not fit inside a /{vpc_prefix} VPC")


def usable(prefix):
    return (1 << (32 - prefix)) - AWS_RESERVED


def allocate_blocks(vpc_net, blocks):
    """VLSM allocator: place each block (largest first) at the next aligned free address.
    Sorting largest-first guarantees alignment never strands space. Mutates blocks with 'cidr'.
    """
    cursor = int(vpc_net.network_address)
    end = int(vpc_net.broadcast_address)
    for b in sorted(blocks, key=lambda x: x["prefix"]):
        size = 1 << (32 - b["prefix"])
        if cursor % size:                      # align up to this block's boundary
            cursor += size - (cursor % size)
        if cursor + size - 1 > end:
            raise ValueError(
                f"VPC {vpc_net} ({vpc_net.num_addresses} addrs) too small to fit "
                f"{b['name']} /{b['prefix']} — shrink IP requests or grow the VPC CIDR")
        b["cidr"] = str(ipaddress.ip_network((cursor, b["prefix"])))
        cursor += size
    return blocks


# ----- planning -------------------------------------------------------------

def build_plan(req):
    azs_default = req.get("azs", ["a", "b", "c"])
    on_prem = [ipaddress.ip_network(x["cidr"]) for x in req.get("external_cidrs", [])]

    # Sanity: chosen LZA supernet (if any) must not collide with on-prem.
    if req.get("supernet"):
        sup = ipaddress.ip_network(req["supernet"])
        hit = [str(o) for o in on_prem if sup.overlaps(o)]
        if hit:
            raise ValueError(f"LZA supernet {sup} overlaps on-prem/peer {hit} — pick another supernet")

    planned_vpcs = []
    all_subnet_nets = []   # to cross-check subnet-vs-subnet and subnet-vs-onprem
    all_vpc_nets = []

    for v in req["vpcs"]:
        vpc_net = ipaddress.ip_network(v["cidr"])
        # VPC vs on-prem
        hit = [str(o) for o in on_prem if vpc_net.overlaps(o)]
        if hit:
            raise ValueError(f"VPC {v['name']} {vpc_net} overlaps on-prem/peer {hit}")
        # VPC vs other VPCs
        clash = [str(n) for n in all_vpc_nets if vpc_net.overlaps(n)]
        if clash:
            raise ValueError(f"VPC {v['name']} {vpc_net} overlaps another VPC {clash}")
        all_vpc_nets.append(vpc_net)

        # expand tiers -> one subnet per AZ
        blocks = []
        for tier in v["tiers"]:
            azs = tier.get("azs", azs_default)
            prefix = prefix_for_ips(tier["ips"], vpc_net.prefixlen)
            for az in azs:
                blocks.append({
                    "name": f"subnet-{tier['name']}-{az}",
                    "tier": tier["name"],
                    "az": az,
                    "prefix": prefix,
                    "ips_requested": tier["ips"],
                    "route_table": tier.get("route_table", f"rt-{tier['name']}"),
                    "type": tier.get("type", "Private"),
                    "purpose": tier.get("purpose", ""),
                })
        allocate_blocks(vpc_net, blocks)

        # subnet vs on-prem + subnet vs all previous subnets
        for b in blocks:
            net = ipaddress.ip_network(b["cidr"])
            hit = [str(o) for o in on_prem if net.overlaps(o)]
            if hit:
                raise ValueError(f"subnet {b['name']} {net} overlaps on-prem {hit}")
            all_subnet_nets.append(net)

        blocks.sort(key=lambda b: int(ipaddress.ip_network(b["cidr"]).network_address))
        planned_vpcs.append({**v, "subnets": blocks})

    return planned_vpcs, on_prem


# ----- excel output ---------------------------------------------------------

HDR = Font(bold=True, color="FFFFFF")
HDR_FILL = PatternFill("solid", fgColor="305496")
TITLE = Font(bold=True, size=13)


def _sheet(wb, name, headers, rows, title=None):
    ws = wb.create_sheet(name)
    r = 1
    if title:
        ws.cell(r, 1, title).font = TITLE
        r += 2
    for c, h in enumerate(headers, 1):
        cell = ws.cell(r, c, h)
        cell.font = HDR
        cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        r += 1
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)
    # autosize-ish
    for c, h in enumerate(headers, 1):
        width = max([len(str(h))] + [len(str(row[c - 1])) for row in rows if c - 1 < len(row)] + [8])
        ws.column_dimensions[get_column_letter(c)].width = min(width + 3, 60)
    ws.freeze_panes = ws.cell(1 + (2 if title else 0) + 1, 1)
    return ws


def write_excel(path, req, planned_vpcs, on_prem):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    customer = req.get("customer", "customer")

    # Accounts
    _sheet(wb, "Accounts",
           ["OU Path", "Account Name", "Account Alias", "Email", "Type", "Description"],
           [[a.get("ou", ""), a.get("name", ""), a.get("alias", a.get("name", "")),
             a.get("email", ""), a.get("type", "Workload"), a.get("description", "")]
            for a in req.get("accounts", [])],
           title=f"{customer} LZA — Accounts")

    # OUs
    _sheet(wb, "OUs",
           ["OU Path", "Parent OU", "Purpose", "ignore"],
           [[o.get("path", ""), o.get("parent", ""), o.get("purpose", ""),
             "true" if o.get("ignore") else ""]
            for o in req.get("ous", [])],
           title=f"{customer} LZA — Organizational Units")

    # VPCs
    _sheet(wb, "VPCs",
           ["Account", "VPC Name", "Region", "VPC CIDR", "VPC Type", "TGW", "#Subnets"],
           [[v.get("account", ""), v["name"], v.get("region", ""), v["cidr"],
             v.get("vpc_type", ""), v.get("tgw", ""), len(v["subnets"])]
            for v in planned_vpcs],
           title=f"{customer} LZA — VPCs")

    # Subnets (the finalized plan)
    _sheet(wb, "Subnets", SUBNET_COLUMNS, list(_subnet_rows(planned_vpcs)),
           title=f"{customer} LZA — Finalized Subnet / CIDR Plan (review me)")

    # External / on-prem CIDRs that were avoided
    _sheet(wb, "External-CIDRs",
           ["Description", "CIDR", "Source"],
           [[x.get("description", ""), x["cidr"], x.get("source", "")]
            for x in req.get("external_cidrs", [])],
           title=f"{customer} — On-Prem / Peer CIDRs AVOIDED by this plan")

    wb.save(path)


# ----- LZA yaml snippet -----------------------------------------------------

def write_yaml_snippet(path, planned_vpcs):
    out = {"vpcs": []}
    for v in planned_vpcs:
        out["vpcs"].append({
            "name": v["name"],
            "account": v.get("account", ""),
            "region": v.get("region", ""),
            "cidrs": [v["cidr"]],
            "subnets": [{
                "name": b["name"],
                "availabilityZone": b["az"],
                "ipv4CidrBlock": b["cidr"],
                "routeTable": b["route_table"],
            } for b in v["subnets"]],
        })
    with open(path, "w") as f:
        f.write("# Generated by plan_subnets.py — paste into network-config.yaml under each VPC.\n")
        f.write("# CIDRs are overlap-checked against on-prem; do not hand-edit without re-validating.\n")
        yaml.safe_dump(out, f, sort_keys=False, default_flow_style=False)


SUBNET_COLUMNS = ["Account", "VPC Name", "Region", "VPC CIDR", "Subnet Name", "Tier", "AZ",
                  "Subnet CIDR", "Prefix", "IPs Requested", "IPs Usable", "Route Table",
                  "Subnet Type", "Purpose"]


def _subnet_rows(planned_vpcs):
    for v in planned_vpcs:
        for b in v["subnets"]:
            yield [v.get("account", ""), v["name"], v.get("region", ""), v["cidr"],
                   b["name"], b["tier"], b["az"], b["cidr"], f"/{b['prefix']}",
                   b["ips_requested"], usable(b["prefix"]), b["route_table"],
                   b["type"], b["purpose"]]


def write_csv(path, planned_vpcs):
    """Flat, diff-friendly CSV of every VPC/subnet/CIDR — the git-reviewable twin of the Excel."""
    with open(path, "w", newline="") as f:
        w = csv.writer(f)
        w.writerow(SUBNET_COLUMNS)
        w.writerows(_subnet_rows(planned_vpcs))


def write_accounts_yaml(path, req):
    """Emit accounts-config.yaml: type 'Mandatory' → mandatoryAccounts, else workloadAccounts.
    LZA expects exactly Management/LogArchive/Audit as mandatory — anything else is a workload."""
    mandatory, workload = [], []
    for a in req.get("accounts", []):
        entry = {
            "name": a["name"],
            "description": a.get("description", ""),
            "email": a["email"],
            "organizationalUnit": a.get("ou", "Root"),
        }
        (mandatory if a.get("type") == "Mandatory" else workload).append(entry)
    doc = {"mandatoryAccounts": mandatory, "workloadAccounts": workload}
    with open(path, "w") as f:
        f.write("# Generated by plan_subnets.py from the intake workbook.\n")
        f.write("# Each email MUST be globally unique across AWS and a real inbox.\n")
        yaml.safe_dump(doc, f, sort_keys=False, default_flow_style=False)


def write_organization_yaml(path, req):
    """Emit organization-config.yaml OU block. Root is implicit (not listed); nested OUs
    use the slash path as the name; parked OUs carry ignore: true."""
    ous = []
    for o in req.get("ous", []):
        path_val = o.get("path", "")
        if path_val in ("", "Root"):           # Root is implicit in LZA
            continue
        entry = {"name": path_val}
        if o.get("ignore"):
            entry["ignore"] = True
        ous.append(entry)
    doc = {"enable": True, "organizationalUnits": ous}
    with open(path, "w") as f:
        f.write("# Generated by plan_subnets.py from the intake workbook.\n")
        f.write("# OU 'name' is the full slash-path; Root is implicit. SCP/tag/backup\n")
        f.write("# policy blocks are added separately in /lza-configure.\n")
        yaml.safe_dump(doc, f, sort_keys=False, default_flow_style=False)


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("requirements", help="path to the requirements YAML")
    ap.add_argument("--out-dir", default=".", help="where to write outputs (default: cwd)")
    args = ap.parse_args()

    with open(args.requirements) as f:
        req = yaml.safe_load(f)

    planned_vpcs, on_prem = build_plan(req)

    os.makedirs(args.out_dir, exist_ok=True)
    customer = req.get("customer", "customer")
    xlsx = os.path.join(args.out_dir, f"{customer}-network-plan.xlsx")
    csvf = os.path.join(args.out_dir, f"{customer}-network-plan.csv")
    snip = os.path.join(args.out_dir, "network-config.snippet.yaml")
    acct = os.path.join(args.out_dir, "accounts-config.yaml")
    org = os.path.join(args.out_dir, "organization-config.yaml")
    write_excel(xlsx, req, planned_vpcs, on_prem)
    write_csv(csvf, planned_vpcs)
    write_yaml_snippet(snip, planned_vpcs)
    write_accounts_yaml(acct, req)
    write_organization_yaml(org, req)

    print(f"✓ Plan validated — no overlap with {len(on_prem)} on-prem/peer range(s).")
    for v in planned_vpcs:
        print(f"  {v['name']:28} {v['cidr']:18} {len(v['subnets'])} subnets")
    n_mand = sum(1 for a in req.get("accounts", []) if a.get("type") == "Mandatory")
    n_work = len(req.get("accounts", [])) - n_mand
    n_ous = sum(1 for o in req.get("ous", []) if o.get("path") not in ("", "Root"))
    print(f"  accounts: {n_mand} mandatory + {n_work} workload   OUs: {n_ous}")
    print(f"\n✓ Review workbook      : {xlsx}")
    print(f"✓ Review CSV (git-diff): {csvf}")
    print(f"✓ accounts-config.yaml : {acct}")
    print(f"✓ organization-config  : {org}")
    print(f"✓ network-config snippet: {snip}")


if __name__ == "__main__":
    main()
